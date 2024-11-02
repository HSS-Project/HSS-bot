import openpyxl
import random

# Excelファイルを読み込む
def get_question_list_from_excel(file_path, start_number, end_number):
    # ワークブックを読み込み
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # 最初のシートを選択

    question_list = []
    #MAXサイズを取得
    if sheet.max_row < end_number:return
    
    for row in range(start_number + 2, end_number + 3):  # Excelの行番号に合わせて+2調整
        quesiton = sheet[f'D{row}'].value  # 単語がD列
        answer = sheet[f'E{row}'].value
        number_list = []
        number_list = random.sample(range(3, sheet.max_row + 1), 3)
        number_list.append(row)
        for _ in range(3):  # 3回シャッフル
            random.shuffle(number_list)
        select = [sheet[f'E{x}'].value for x in number_list]
        if quesiton and answer:
            dict_data = {
                "question":quesiton,
                "answer":answer,
                "select":select
            }
            question_list.append(dict_data)

    return question_list

# 使用例
file_path = 'DataBase4500.xlsx'  # 読み込むxlsxファイルのパスを指定
start_number = 1  # 開始番号
end_number = 5    # 終了番号
question_list = get_question_list_from_excel(file_path, start_number, end_number)
print(question_list[2])