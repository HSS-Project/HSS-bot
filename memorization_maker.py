from typing import TypedDict, Literal
from typing_extensions import NotRequired

import json
import random
import openpyxl
from HSS import User
import discord

from typing import TypedDict


class QuestionData(TypedDict):
    question: str
    mode: int
    answer: str
    select: NotRequired[list[str]]


class CardData(TypedDict):
    questions: list[QuestionData]
    sharecode: int


class StatusData(TypedDict):
    count: int
    score: int

class answerTypedDict(TypedDict):
    answer: str

class MemorizationData(answerTypedDict):
    memorization: dict[str, dict[str, CardData]]
    user_status: dict[str, dict[str, StatusData]]


class MemorizationSystem:
    data: MemorizationData
    

    def __init__(self, filename: str = 'memorization.json'):
        self.filename = filename
        self.data = {"memorization": {}, "user_status": {}}

    async def checkuser_in_HSS(self, interaction:discord.Interaction) -> bool:
        """
        Check if the user is in the HSS.
        """
        with open("token.json", "r", encoding="utf-8") as f:
            token = json.load(f)
            token = token["HSSAPI_TOKEN"]
        user = User(token=token)
        try:
            user.get_permission_discordUserID(str(interaction.user.id))
        except:
            await interaction.response.send_message("この機能はHSSにログインして学校に登録していないユーザーはご利用いただけません。\n[HSSにloginする](https://hss.aknet.tech/login)", ephemeral=True)
            return False
        return True

    async def save_data(self):
        """
        Save the data to a file in JSON format.

        Args:
            None

        Returns:
            None
        """
        with open(self.filename, 'w', encoding='utf-8') as file:
            json.dump(self.data, file, ensure_ascii=False, indent=2)

    async def load_data(self):
        """
        Loads the data from a file and save it for `self.data`.

        Returns:
            None
        """
        try:
            with open(self.filename, 'r', encoding='utf-8') as file:
                self.data = json.load(file)
        except FileNotFoundError:
            self.data = {"memorization": {}, "user_status": {}}

    async def del_mission_title(self, id_: str, title: str):
        """
        Delete a mission title from the memorization data.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.

        Returns:
            bool: True if the title is deleted successfully, False otherwise.
        """
        num_id = str(id_)
        await self.load_data()
        if num_id in self.data["memorization"]:
            if title in self.data["memorization"][num_id]:
                self.data["memorization"][num_id].pop(title)
                await self.save_data()
                return True
        return False


    async def add_mission(self, id_: str, title: str,random_number:int,mode: int, mission: str, answer: str, select: list | None = None, ):
        """
        Add a mission to the memorization data.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.
            random_number (int): The share code of the mission.
            mode (int): The mode of the mission. 0 for normal mission, 1 for multiple-choice mission.
            mission (str): The mission question.
            answer (str): The answer to the mission.
            select (list, optional): The list of options for multiple-choice mission. Defaults to None.

        Returns:
            bool: True if the mission is added successfully, False otherwise.
        """
        num_id = str(id_)
        await self.load_data()
        self.data["memorization"].setdefault(num_id, {})
        self.data["memorization"][num_id].setdefault(title, {"questions": [], "sharecode": random_number})
        self.data["memorization"][num_id][title]["questions"].append({"question": mission, "mode": mode, "answer": answer})
        if mode == 1:
            assert isinstance(select, list)
            self.data["memorization"][num_id][title]["questions"][-1]["select"] = select
        await self.save_data()
        return True

    async def add_mission_Excel(self,id_: str, title: str, number: int, workbook: openpyxl.Workbook):
        """
        Add a mission to the memorization data from an Excel file.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.
            number (int): The share code of the mission.
            workbook (openpyxl.Workbook): The workbook of the mission.
        
        Returns:
            bool: True if the mission is added successfully, False otherwise.
        """
        num_id = str(id_)
        await self.load_data()
        sheet = workbook.active
        assert sheet is not None
        self.data["memorization"].setdefault(num_id, {})
        self.data["memorization"][num_id].setdefault(title, {"questions": [], "sharecode": number})
        row_count = sum(1 for _ in sheet.iter_rows(min_row=1, values_only=True))
        if row_count < 4:return False
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row[0] or not row[1] or not row[2]:continue
            mode = int(row[2])
            if not isinstance(mode, int):return False
            question = row[0]
            answer = row[1]
            if mode == 1:
                self.data["memorization"][num_id][title]["questions"].append({"question": question, "mode": 0, "answer": answer})
            elif mode == 2:
                select = []
                try:
                    if row[3] is not None and row[4] is not None and row[5] is not None and row[6] is not None:
                        select = [row[3], row[4], row[5], row[6]]
                        answer_num = select.index(answer) + 1
                        mode_change = 0
                    else:
                        mode_change = 1
                except:
                    mode_change = 1
                if mode_change == 1:
                    num_rows = sheet.max_row 
                    random_answer_index = random.randint(0, 3)
                    select = []
                    for _ in range(4):
                        while True:
                            random_select = random.randint(1, num_rows + 1) 
                            cell = sheet.cell(row=random_select, column=2)
                            if cell.value not in select and cell.value is not None:
                                if not cell.value  == answer:
                                    select.append(cell.value)
                                    break
                    select[random_answer_index] = answer
                    answer_num =  select.index(answer)+1
                self.data["memorization"][num_id][title]["questions"].append({"question": question, "mode": 1, "answer": answer_num, "select": select})
        await self.save_data()
        return True
                
    async def del_mission(self, id_: str, title: str, question: str):
        """
        Delete a mission from the memorization data.

        Parameters:
        - id (int): The User ID.
        - title (str): The title of the memorization data.
        - question (str): The question to be deleted.

        Returns:
        - bool: True if the question is successfully deleted, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"]:
            for cont, item in enumerate(self.data["memorization"][num_id][title]["questions"]):
                if item["question"] == question:
                    self.data["memorization"][num_id][title]["questions"].pop(cont)
                    await self.save_data()
                    return True
        return False

    async def edit_misson(self, id_: str, title: str, number: int, modes: int, value: str, select_number: int | None = None):
        """
        Edit a mission in the memorization data.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.
            number (int): The number of the content to edit.
            mode (str): The mode of the content to edit.
            answer (str): The new answer for the content.
            select_number (int, optional): The number of the select option to edit (only applicable if mode is "select"). Defaults to None.
            modes (int): The mode of the mission. 0 for question mission, 1 for answer 2 for select_answer.
        Returns:
            bool: True if the mission was successfully edited, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"] and title in self.data["memorization"][num_id]:
            edit_before = self.data["memorization"][num_id][title]["questions"][number]
            if edit_before["mode"] == 0:
                if modes == 0:
                    edit_before["question"] = value
                elif modes == 1:
                    edit_before["answer"] = value
            elif edit_before["mode"] == 1:
                if modes == 0:
                    edit_before["question"] = value
                elif modes == 1:
                    edit_before["answer"] = value
                elif modes == 2:
                    assert select_number is not None and "select" in edit_before
                    if not value in edit_before["select"]:
                        edit_before["select"][select_number] = value
                    else:
                        return False
            self.data["memorization"][num_id][title]["questions"][number] = edit_before
            await self.save_data()
            return True
        return False

    async def get_mission(self, id_: str, title: str) -> Literal[False] | CardData:
        """
        Get the content of a mission based on its ID and title.

        Parameters:
        - id (str): The ID of the mission.
        - title (str): The title of the mission.

        Returns:
        - content (list): The content of the mission if it exists, False otherwise.
        """
        await self.load_data()
        return self.data["memorization"].get(id_, {}).get(title, False)

    async def make_sharecode(self) -> int:
        """
        Make a sharecode for a mission.

        Returns:
        - random_number (int): The sharecode for the mission.
        """
        await self.load_data()
        while True:
            random_number = random.randint(10000000, 99999999)
            for num_id in self.data["memorization"]:
                for title in self.data["memorization"][num_id]:
                    if self.data["memorization"][num_id][title]["sharecode"] == random_number:
                        continue
            return random_number

    async def get_sharecode(self, id_: str, title: str) -> Literal[False] | int:
        """
        Get the sharecode of a mission.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.

        Returns:
            int: The sharecode of the mission if it exists, False otherwise.
        """
        num_id = str(id_)
        await self.load_data()
        if num_id in self.data["memorization"] and title in self.data["memorization"][num_id]:
            return self.data["memorization"][num_id][title]["sharecode"]
        return False

    async def get_mission_sharecode(self, code) -> Literal[False] | list[QuestionData]:
        """
        Get the content of a mission based on its sharecode.

        Parameters:
        - code (int): The sharecode of the mission.

        Returns:
        - content (list): The content of the mission if it exists, False otherwise.
        """
        await self.load_data()
        for num_id in self.data["memorization"]:
            for title in self.data["memorization"][num_id]:
                if self.data["memorization"][num_id][title]["sharecode"] == code:
                    return self.data["memorization"][num_id][title]["questions"]
        return False 
    
    async def sharecode_question_copy(self, id_: str, sharecode: int):
        """
        Copy a mission from a sharecode.

        Args:
            id (str): The ID of the mission.
            sharecode (int): The sharecode of the mission.
        
        Returns:
            bool: True if the mission was copied successfully, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        self.data["memorization"].setdefault(num_id, {})
        await self.save_data()#消したらなぜか動かないww
        await self.load_data()
        for bef_id in self.data["memorization"]:
            for title in self.data["memorization"][bef_id]:
                if self.data["memorization"][bef_id][title].get("sharecode") == sharecode:
                    self.data["memorization"][num_id][title] = CardData(
                        questions=self.data["memorization"][bef_id][title]["questions"],
                        sharecode=await self.make_sharecode()
                    )
                    await self.save_data()
                    return True
        return False

    async def sharecode_true(self, id_: str, title: str) -> bool:
        """
        Check if the mission has a sharecode.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.

        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"]:
            try:
                code = self.data["memorization"][num_id][title]["sharecode"]
            except:
                return False
            return bool(code)
        return False

    async def get_mission_title(self, id_: str):
        """
        Get the mission titles for a given ID.

        Args:
            id (str): The ID of the mission.

        Returns:
            list or bool: A list of mission titles if the ID exists in the data, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"]:
            return list(self.data["memorization"][num_id].keys())
        return False

    async def check_answer(self, id_: str, title: str, question: str, answer: str, mode: int) -> bool:
        """
        Check the answer to a mission.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
            question (str): The question of the mission.
            answer (str): The answer to the mission.
            mode (int): The mode of the mission. 0 for normal mission, 1 for multiple-choice mission.

        Returns:
            bool: True if the answer is correct, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"] and title in self.data["memorization"][num_id]:
            for item in self.data["memorization"][num_id][title]["questions"]:
                if item["question"] == question:
                    if mode == 0:
                        return item["answer"] == answer
                    else:
                        return answer == item["select"][int(item["answer"])-1]
        return False
    
    async def get_answer(self, id_: str, title: str, question: str):
        """
        Get the answer to a mission.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
            question (str): The question of the mission.

        Returns:
            str: The answer to the mission if it exists, False otherwise.
        """
        num_id = str(id_)
        await self.load_data()
        if num_id in self.data["memorization"] and title in self.data["memorization"][num_id]:
            for item in self.data["memorization"][num_id][title]["questions"]:
                if item["question"] == question:
                    if item["mode"] == 0:
                        return item["answer"]
                    elif item["mode"] == 1:
                        return item["select"][int(item["answer"])-1]
                                            
        return False
    
    async def randam_mission_select(self,id_:str,title:str):
        """
        Randomly select a mission.
        
        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
        
        Returns:
            bool: True if the mission was successfully selected, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"] and title in self.data["memorization"][num_id]:
            random.shuffle(self.data["memorization"][num_id][title]["questions"])
            
            await self.save_data()
            return True
        return False

    async def get_mission_selectmode_list(self,id_:str,title:str):
        """
        Get the mode of the mission.
        
        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
            question_number (int): The number of the question.
            
        Returns:
            list: The mode of the mission.
        """
        await self.load_data()
        num_id = str(id_)
        selects = []
        if num_id in self.data["memorization"] and title in self.data["memorization"][num_id]:
            for item in self.data["memorization"][num_id][title]["questions"]:
                if item["mode"] == 1:
                    selects.append(item["mode"])
            await self.save_data()
            return selects

    async def select_question_randam(self,id_:str,title:str,question_number:int):
        """
        Randomly select a question.
        
        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
            question_number (int): The number of the question.
            
        Returns:
            bool: True if the question was successfully selected, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        anwer_num = int(self.data["memorization"][num_id][title]["questions"][question_number]["answer"])-1
        selects:list = self.data["memorization"][num_id][title]["questions"][question_number]["select"]
        answer = selects[anwer_num]
        random.shuffle(selects)
        answer_number = selects.index(answer)+1
        self.data["memorization"][num_id][title]["questions"][question_number]["select"] = selects
        self.data["memorization"][num_id][title]["questions"][question_number]["answer"] = answer_number
        await self.save_data()
        return True

    async def memorization_sheet(self,id_:str, title:str):
        """
        Get the memorization sheet.
        
        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
            
        Returns:
            str:sheet_text 
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["memorization"] and (title in self.data["memorization"][num_id]):
            make_sheet = ""
            for item in self.data["memorization"][num_id][title]["questions"]:
                question_text = item["question"]
                if item["mode"] == 0:
                    answer_text = item["answer"]
                elif item["mode"] == 1:
                    answer_text = item["select"][item["answer"]-1]
                make_sheet += f"{question_text} : ||{answer_text}||\n"
                if len(make_sheet) > 1900:
                    make_sheet += "......\n"
                    return make_sheet
            return make_sheet     
        return False                       
            
    """
    user_status System ↓
    """

    async def add_user_status(self, id_: str, title: str) -> Literal[True]:
        """
        Add a user status to the user status data.

        Args:
            id (str): The ID of the user.
            title (str): The title of the mission.

        Returns:
            bool: True if the user status is added successfully, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        self.data["user_status"].setdefault(num_id, {title: {"count": 0, "score": 0}})
        self.data["user_status"][num_id].setdefault(title, {"count": 0, "score": 0})
        await self.save_data()
        return True

    async def edit_user_status(self, id_: str, title: str, count: int, score: int) -> bool:
        """
        Edit a user status in the user status data.

        Args:
            id (str): The ID of the user.
            title (str): The title of the mission.
            count (int): The count of the user status.
            score (int): The score of the user status.
        
        Returns:
            bool: True if the user status was successfully edited, False otherwise.
        """
        await self.load_data()
        num_id = str(id_)
        if num_id in self.data["user_status"] and title in self.data["user_status"][num_id]:
            edit_before = self.data["user_status"][num_id][title]
            edit_before["count"] += count
            edit_before["score"] += score
            self.data["user_status"][num_id][title] = edit_before
            await self.save_data()
            return True
        return False

    async def get_user_status(self, id_: str, title: str) -> StatusData | Literal[False]:
        """
        Get the user status of a user based on its ID and title.

        Args:
            id (str): The ID of the user.
            title (str): The title of the mission.

        Returns:
            dict: The user status of the user if it exists, False otherwise.
        """
        num_id = str(id_)
        await self.load_data()
        return self.data["user_status"].get(num_id, {}).get(title, False)