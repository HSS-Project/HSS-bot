import memorization_maker.share as share
import memorization_maker.genre as genre
import memorization_maker.base_question_add as maker_add
import openpyxl
import random

class Vocabulary:
    def __init__(self) -> None:
        self.share = share.Share()
        self.genre = genre.Genre()
        self.make_add = maker_add.Add()
        self.xlsx_file_path = "DataBase4500.xlsx"

    async def make_vocabulary(self,user_id:str,title:str,start_number:int,end_unmber:int):
        num_id = str(user_id)
        sharecode = await self.share.make_sharecode()
        make_title = f"{title}_{start_number}-{end_unmber}"
        await self.make_add.init_add(str(user_id),make_title,sharecode)       
        await self.genre.make_genre(num_id,"vocabulary")
        await self.genre.add_genre(num_id,"vocabulary",sharecode)
        list_data = await self.get_question_list_from_excel(start_number,end_unmber)
        for data in list_data:await self.make_add.add_misson_select(sharecode,data["question"],data["answer"],data["select"])
        return sharecode

    async def get_question_list_from_excel(self,start_number, end_number):
        wb = openpyxl.load_workbook(self.xlsx_file_path)
        sheet = wb.active
        question_list = []
        if sheet.max_row < end_number:return
        for row in range(start_number + 2, end_number + 3):
            quesiton = sheet[f'D{row}'].value
            answer = sheet[f'E{row}'].value
            number_list = []
            number_list = random.sample(range(3, sheet.max_row + 1), 3)
            number_list.append(row)
            for _ in range(3):random.shuffle(number_list)
            select = [sheet[f'E{x}'].value for x in number_list]
            answer_index = select.index(answer)
            if quesiton and answer:
                dict_data = {"question":quesiton,"answer":answer_index,"select":select}
                question_list.append(dict_data)

        return question_list