import Read_and_Write
import owner_manager
import re
from typing import TypedDict, Literal
from typing_extensions import NotRequired
from typing import Union
import openpyxl
import random

class QuestionData(TypedDict):
    question: str
    answer: str
    mode: int
    select: NotRequired[list[str]]

class CardData(TypedDict):
    questions: list[QuestionData]
    onwer: list[str]
    sharecode: Union[int, bool]

class answerTypedDict(TypedDict):
    answer: str

class MemorizationData(answerTypedDict):
    memorization: dict[dict[str, CardData]]

class Add:
    def __init__(self):
        self.rw = Read_and_Write.Read_and_Write()
        self.owner = owner_manager.OwnerManager()
        self.base_data:dict = {"memorization":{}, "genre":{}}

    async def init_add(self,user_id:str,title:str,_sharecode):
        num_id = str(user_id)
        sharecode = str(_sharecode)
        self.base_data:dict = await self.rw.load_base()
        self.base_data["memorization"][num_id].setdefault(sharecode, {"title":f"{title}_{sharecode}","questions": [], "onwer":[num_id]})
        return self.base_data        

    async def add_misson(self, _id:str,title,_sharecode,quetion:str,answer:str):
        user_id = str(_id)
        sharecode = str(_sharecode)
        self.base_data:dict = await self.init_add(user_id,title,sharecode)
        if not await self.owner.ower_check(user_id,title):return False
        self.base_data["memorization"][sharecode]["questions"].append({"question":quetion,"answer":answer,"mode":0})
        await self.rw.write_base(self.base_data)
        
    async def add_misson_select(self,_id:str,title,_sharecode,quetion:str,answer:str,select:list):
        user_id = str(_id)
        sharecode = str(_sharecode)
        self.base_data:dict = await self.init_add(user_id,title,sharecode)
        if not await self.owner.ower_check(user_id,title):return False
        select.append(answer)
        select = random.sample(select, len(select))
        self.base_data["memorization"][sharecode]["questions"].append({"question":quetion,"answer":select.index(answer)+1,"mode":1,"select":select})
        await self.rw.write_base(self.base_data)
        self.base_data["memorization"][sharecode]["questions"].append({"question":quetion,"answer":answer,"mode":1,"select":select})
        await self.rw.write_base(self.base_data)

    async def replace_parentheses(self,text):
        # ()内のテキストを抽出
        answers:list = re.findall(r'\((.*?)\)', text)
        
        # ()内のテキストを①、②、③...に置き換え
        if len(answers) > 5:
            return False, False
        for i, answer in enumerate(answers):
            number = chr(9312 + i)  # 9312はUnicodeで①の値
            text:str = text.replace(f'({answer})', f'{number}')
        return text, answers
    
    def replace_numbers_with_answers(self,text, answers):
        # 正規表現で番号を探して置き換える
        for i, answer in enumerate(answers):
            number = chr(9312 + i)  # 9312はUnicodeで①の値
            text = text.replace(f'{number}', f'||{answer}||')
        return text
    
    async def add_misson_text(self,_id:str,title,_sharecode,text:str):
        user_id = str(_id)
        sharecode = str(_sharecode)
        self.base_data:dict = await self.init_add(user_id,title,sharecode)
        if not await self.owner.ower_check(user_id,title):return False
        text, answers = await self.replace_parentheses(text)
        if text is False or answers is False:return False
        self.base_data["memorization"][sharecode]["questions"].append({"question":text,"answer":answers, "mode":2})
        await self.rw.write_base(self.base_data)
        
    async def add_misson_in_Excel(self,user_id:str,title,_sharecode,text:str,excelfile:openpyxl.Workbook):
        sharecode = str(_sharecode)
        self.base_data:dict = await self.init_add(user_id,title,sharecode)
        if not await self.owner.ower_check(user_id,title):return False
        sheet = excelfile.active
        row_count = sum(1 for _ in sheet.iter_rows(min_col=1,values_only=True))
        if row_count > 4:return False
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0] or row[1] or row[2]:
                mode = int(row[2])
                if not isinstance(mode, int):return False
                question = row[0]
                answer = row[1]
                if mode == 0:
                    self.base_data["memorization"][sharecode]["questions"].append({"question":question,"answer":answer,"mode":0})
                elif mode == 1:
                    select = []
                    if row[3] or row[4] or row[5] or row[6]:
                        select = [row[3],row[4],row[5],row[6]]
                        answer_num = select.index(answer) + 1
                    else:
                        num_rows_max = sheet.max_row
                        random_answer_index = random.sample(range(3,num_rows_max), 3)
                        for _ in range(4):
                            while True:
                                random_select = random.randint(1, num_rows_max + 1) 
                                cell = sheet.cell(row=random_select, column=2)
                                if cell.value not in select and cell.value is not None:
                                    if not cell.value  == answer:
                                        select.append(cell.value)
                                        break
                        select[random_answer_index] = answer
                        answer_num = select.index(answer) + 1
                    self.base_data["memorization"][sharecode]["questions"].append({"question":question,"answer":answer_num,"mode":1,"select":select})
            elif row[0] and int(row[2]) == 2:
                text, answers = await self.replace_parentheses(text)
                if text is False or answers is False:return False
                self.base_data["memorization"][sharecode]["questions"].append({"question":text,"answer":answers, "mode":2})
            
        await self.rw.write_base(self.base_data)
        return True