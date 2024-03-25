from typing import TypedDict, Literal
from typing_extensions import NotRequired

import json
import random
import openpyxl


class QuestionData(TypedDict):
    question: str
    mode: int
    answer: str
    select: NotRequired[list[str]]


class CardData(TypedDict):
    questions: list[QuestionData]
    sharecode: int


class StatusData(TypedDict):
    count: int
    score: int


class MemorizationData(TypedDict):
    memorization: dict[str, dict[str, CardData]]
    user_status: dict[str, dict[str, StatusData]]


class MemorizationSystem:
    data: MemorizationData

    def __init__(self, filename: str = 'memorization.json'):
        self.filename = filename
        self.data = {"memorization": {}, "user_status": {}}

    async def save_data(self):
        """
        Save the data to a file in JSON format.

        Args:
            None

        Returns:
            None
        """
        with open(self.filename, 'w', encoding='utf-8') as file:
            json.dump(self.data, file, ensure_ascii=False, indent=2)

    async def load_data(self):
        """
        Loads the data from a file and save it for `self.data`.

        Returns:
            None
        """
        try:
            with open(self.filename, 'r', encoding='utf-8') as file:
                self.data = json.load(file)
        except FileNotFoundError:
            self.data = {"memorization": {}, "user_status": {}}

    async def del_mission_title(self, id: str, title: str):
        """
        Delete a mission title from the memorization data.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.

        Returns:
            bool: True if the title is deleted successfully, False otherwise.
        """
        await self.load_data()
        if id in self.data["memorization"]:
            if title in self.data["memorization"][id]:
                self.data["memorization"][id].pop(title)
                await self.save_data()
                return True
        return False


    async def add_mission(self, id: str, title: str,random_number:int,mode: int, mission: str, answer: str, select: list | None = None, ):
        """
        Add a mission to the memorization data.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.
            random_number (int): The share code of the mission.
            mode (int): The mode of the mission. 0 for normal mission, 1 for multiple-choice mission.
            mission (str): The mission question.
            answer (str): The answer to the mission.
            select (list, optional): The list of options for multiple-choice mission. Defaults to None.

        Returns:
            bool: True if the mission is added successfully, False otherwise.
        """
        id = str(id)
        await self.load_data()
        self.data["memorization"].setdefault(id, {})
        self.data["memorization"][id].setdefault(title, {"questions": [], "sharecode": random_number})
        self.data["memorization"][id][title]["questions"].append({"question": mission, "mode": mode, "answer": answer})
        if mode == 1:
            assert isinstance(select, list)
            self.data["memorization"][id][title]["questions"][-1]["select"] = select
        await self.save_data()
        return True

    async def add_mission_Excel(self,id: str, title: str, number: int, workbook: openpyxl.Workbook):
        """
        Add a mission to the memorization data from an Excel file.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.
            number (int): The share code of the mission.
            workbook (openpyxl.Workbook): The workbook of the mission.
        
        Returns:
            bool: True if the mission is added successfully, False otherwise.
        """
        id = str(id)
        await self.load_data()
        sheet = workbook.active
        assert sheet is not None
        self.data["memorization"].setdefault(id, {})
        self.data["memorization"][id].setdefault(title, {"questions": [], "sharecode": number})
        row_count = sum(1 for _ in sheet.iter_rows(min_row=1, values_only=True))
        ch = 0
        if row_count < 4:
            ch = 1

        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row[0] or not row[1] or not row[2]:
                continue

            question = row[0]
            answer = row[1]
            mode:int = row[2]
            if mode == 1:
                self.data["memorization"][id][title]["questions"].append({"question": question, "mode": 0, "answer": answer})
            elif mode == 2:
                select = []
                if row[3] is not None and row[4] is not None and row[5] is not None and row[6] is not None:
                    select = [row[3], row[4], row[5], row[6]]
                    random_answer = select.index(answer) + 1
                else:
                    if ch == 1:
                        return False
                    random_selects = random.sample(range(1, 5), 4)
                    for random_select in random_selects:
                        cell = sheet.cell(row=random_select, column=2)
                        select.append(cell.value)
                        random_answer = select.index(answer) + 1 if answer in select else random.randint(1, 4)
                    select[random_answer - 1] = answer

                self.data["memorization"][id][title]["questions"].append({"question": question, "mode": 1, "answer": random_answer, "select": select})
        await self.save_data()
        return True
                
    async def del_mission(self, id: str, title: str, question: str):
        """
        Delete a mission from the memorization data.

        Parameters:
        - id (int): The User ID.
        - title (str): The title of the memorization data.
        - question (str): The question to be deleted.

        Returns:
        - bool: True if the question is successfully deleted, False otherwise.
        """
        await self.load_data()
        if id in self.data["memorization"]:
            for cont, item in enumerate(self.data["memorization"][id][title]["questions"]):
                if item["question"] == question:
                    self.data["memorization"][id][title]["questions"].pop(cont)
                    await self.save_data()
                    return True
        return False

    async def edit_misson(self, id: str, title: str, number: int, modes: int, value: str, select_number: int | None = None):
        """
        Edit a mission in the memorization data.

        Args:
            id (str): The User ID.
            title (str): The title of the mission.
            number (int): The number of the content to edit.
            mode (str): The mode of the content to edit.
            answer (str): The new answer for the content.
            select_number (int, optional): The number of the select option to edit (only applicable if mode is "select"). Defaults to None.
            modes (int): The mode of the mission. 0 for question mission, 1 for answer 2 for select_answer.
        Returns:
            bool: True if the mission was successfully edited, False otherwise.
        """
        await self.load_data()
        if id in self.data["memorization"] and title in self.data["memorization"][id]:
            edit_before = self.data["memorization"][id][title]["questions"][number]
            if edit_before["mode"] == 0:
                if modes == 0:
                    edit_before["question"] = value
                elif modes == 1:
                    edit_before["answer"] = value
            elif edit_before["mode"] == 1:
                if modes == 0:
                    edit_before["question"] = value
                elif modes == 1:
                    edit_before["answer"] = value
                elif modes == 2:
                    assert select_number is not None and "select" in edit_before
                    edit_before["select"][select_number] = value

            self.data["memorization"][id][title]["questions"][number] = edit_before
            await self.save_data()
            return True
        return False

    async def get_mission(self, id: str, title: str) -> Literal[False] | CardData:
        """
        Get the content of a mission based on its ID and title.

        Parameters:
        - id (str): The ID of the mission.
        - title (str): The title of the mission.

        Returns:
        - content (list): The content of the mission if it exists, False otherwise.
        """
        await self.load_data()
        return self.data["memorization"].get(id, {}).get(title, False)

    async def make_sharecode(self) -> int:
        """
        Make a sharecode for a mission.

        Returns:
        - random_number (int): The sharecode for the mission.
        """
        await self.load_data()
        while True:
            random_number = random.randint(10000000, 99999999)
            for id in self.data["memorization"]:
                for title in self.data["memorization"][id]:
                    if self.data["memorization"][id][title]["sharecode"] == random_number:
                        continue
            return random_number

    async def get_sharecode(self, id: str, title: str) -> Literal[False] | int:
        """
        Get the sharecode of a mission.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.

        Returns:
            int: The sharecode of the mission if it exists, False otherwise.
        """
        await self.load_data()
        if id in self.data["memorization"] and title in self.data["memorization"][id]:
            return self.data["memorization"][id][title]["sharecode"]
        return False

    async def get_mission_sharecode(self, code) -> Literal[False] | list[QuestionData]:
        """
        Get the content of a mission based on its sharecode.

        Parameters:
        - code (int): The sharecode of the mission.

        Returns:
        - content (list): The content of the mission if it exists, False otherwise.
        """
        await self.load_data()
        for id in self.data["memorization"]:
            for title in self.data["memorization"][id]:
                if self.data["memorization"][id][title]["sharecode"] == code:
                    return self.data["memorization"][id][title]["questions"]
        return False 
    
    async def sharecode_question_copy(self, id: str, sharecode: int):
        await self.load_data()
        id = str(id)
        self.data["memorization"].setdefault(id, {})
        await self.save_data()#消したらなぜか動かないww
        await self.load_data()
        for bef_id in self.data["memorization"]:
            for title in self.data["memorization"][bef_id]:
                if self.data["memorization"][bef_id][title].get("sharecode") == sharecode:
                    self.data["memorization"][id][title] = CardData(
                        questions=self.data["memorization"][bef_id][title]["questions"],
                        sharecode=await self.make_sharecode()
                    )
                    await self.save_data()
                    return True
        return False

    async def sharecode_true(self, id: str, title: str) -> bool:
        """
        Check if the mission has a sharecode.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.

        """
        await self.load_data()
        if id in self.data["memorization"]:
            try:
                code = self.data["memorization"][id][title]["sharecode"]
            except:
                return False
            return bool(code)
        return False

    async def get_mission_title(self, id: str):
        """
        Get the mission titles for a given ID.

        Args:
            id (str): The ID of the mission.

        Returns:
            list or bool: A list of mission titles if the ID exists in the data, False otherwise.
        """
        await self.load_data()
        if id in self.data["memorization"]:
            return list(self.data["memorization"][id].keys())
        return False

    async def check_answer(self, id: str, title: str, question: str, answer: str, mode: int) -> bool:
        """
        Check the answer to a mission.

        Args:
            id (str): The ID of the mission.
            title (str): The title of the mission.
            question (str): The question of the mission.
            answer (str): The answer to the mission.
            mode (int): The mode of the mission. 0 for normal mission, 1 for multiple-choice mission.

        Returns:
            bool: True if the answer is correct, False otherwise.
        """
        await self.load_data()
        if id in self.data["memorization"] and title in self.data["memorization"][id]:
            for item in self.data["memorization"][id][title]["questions"]:
                if item["question"] == question:
                    if mode == 0:
                        return item["answer"] == answer
                    else:
                        assert "select" in item
                        return answer == item["select"][int(item["answer"])-1]
        return False

    """
    user_status System ↓
    """

    async def add_user_status(self, id: str, title: str) -> Literal[True]:
        """
        Add a user status to the user status data.

        Args:
            id (str): The ID of the user.
            title (str): The title of the mission.

        Returns:
            bool: True if the user status is added successfully, False otherwise.
        """
        await self.load_data()
        self.data["user_status"].setdefault(id, {title: {"count": 0, "score": 0}})
        self.data["user_status"][id].setdefault(title, {"count": 0, "score": 0})
        await self.save_data()
        return True

    async def edit_user_status(self, id: str, title: str, count: int, score: int) -> bool:
        """
        Edit a user status in the user status data.

        Args:
            id (str): The ID of the user.
            title (str): The title of the mission.
            count (int): The count of the user status.
            score (int): The score of the user status.
        
        Returns:
            bool: True if the user status was successfully edited, False otherwise.
        """
        await self.load_data()
        if id in self.data["user_status"] and title in self.data["user_status"][id]:
            edit_before = self.data["user_status"][id][title]
            edit_before["count"] += count
            edit_before["score"] += score
            self.data["user_status"][id][title] = edit_before
            await self.save_data()
            return True
        return False

    async def get_user_status(self, id: str, title: str) -> StatusData | Literal[False]:
        """
        Get the user status of a user based on its ID and title.

        Args:
            id (str): The ID of the user.
            title (str): The title of the mission.

        Returns:
            dict: The user status of the user if it exists, False otherwise.
        """
        await self.load_data()
        return self.data["user_status"].get(id, {}).get(title, False)
